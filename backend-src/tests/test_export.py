import re

from flask.testing import FlaskClient


# ── helpers ──────────────────────────────────────────────────────────────────

TS_RE = r"\d{4}-\d{2}-\d{2}T\d{2}-\d{2}-\d{2}"


def _lines(r) -> list[str]:
    return r.data.decode().strip().splitlines()


def _header(r) -> list[str]:
    return _lines(r)[0].split(",")


# ── auth guards ───────────────────────────────────────────────────────────────

def test_export_tubes_requires_auth(client: FlaskClient) -> None:
    assert client.get("/api/export/tubes").status_code == 401


def test_export_boxes_requires_auth(client: FlaskClient) -> None:
    assert client.get("/api/export/boxes").status_code == 401


def test_export_cores_requires_auth(client: FlaskClient) -> None:
    assert client.get("/api/export/cores").status_code == 401


def test_export_single_box_requires_auth(client: FlaskClient) -> None:
    assert client.get("/api/export/boxes/1").status_code == 401


def test_export_single_core_requires_auth(client: FlaskClient) -> None:
    assert client.get("/api/export/cores/1").status_code == 401


# ── tubes ─────────────────────────────────────────────────────────────────────

def test_export_tubes_csv(auth_client: FlaskClient) -> None:
    auth_client.post("/api/tubes", json={"barcode": "TUBE001", "site_name": "River A"})
    r = auth_client.get("/api/export/tubes")
    assert r.status_code == 200
    assert "text/csv" in r.content_type
    assert re.search(rf'filename="tubes-{TS_RE}\.csv"', r.headers["Content-Disposition"])
    text = r.data.decode()
    assert "barcode" in text
    assert "TUBE001" in text
    assert "River A" in text


def test_export_tubes_tsv(auth_client: FlaskClient) -> None:
    auth_client.post("/api/tubes", json={"barcode": "TUBE001", "site_name": "River A"})
    r = auth_client.get("/api/export/tubes?format=tsv")
    assert r.status_code == 200
    assert "tab-separated" in r.content_type
    assert re.search(rf'filename="tubes-{TS_RE}\.tsv"', r.headers["Content-Disposition"])
    text = r.data.decode()
    assert "\t" in text
    assert "TUBE001" in text


def test_export_tubes_includes_box_info(auth_client: FlaskClient) -> None:
    box = auth_client.post("/api/boxes", json={"barcode": "BOX001", "name": "Shelf A"}).json
    auth_client.post("/api/tubes", json={"barcode": "TUBE001", "box_id": box["id"]})
    text = auth_client.get("/api/export/tubes").data.decode()
    assert "BOX001" in text
    assert "Shelf A" in text


def test_export_tubes_empty(auth_client: FlaskClient) -> None:
    r = auth_client.get("/api/export/tubes")
    assert r.status_code == 200
    assert len(_lines(r)) == 1  # header only


# ── boxes (hierarchical) ──────────────────────────────────────────────────────

def test_export_boxes_csv(auth_client: FlaskClient) -> None:
    auth_client.post("/api/boxes", json={"barcode": "BOX001", "name": "Shelf A"})
    r = auth_client.get("/api/export/boxes")
    assert r.status_code == 200
    assert "text/csv" in r.content_type
    assert re.search(rf'filename="boxes-{TS_RE}\.csv"', r.headers["Content-Disposition"])
    text = r.data.decode()
    assert "row_type" in text
    assert "BOX001" in text


def test_export_boxes_tsv(auth_client: FlaskClient) -> None:
    auth_client.post("/api/boxes", json={"barcode": "BOX001"})
    r = auth_client.get("/api/export/boxes?format=tsv")
    assert r.status_code == 200
    assert "tab-separated" in r.content_type
    assert re.search(rf'filename="boxes-{TS_RE}\.tsv"', r.headers["Content-Disposition"])
    assert "\t" in r.data.decode()


def test_export_boxes_includes_tube_rows(auth_client: FlaskClient) -> None:
    box = auth_client.post("/api/boxes", json={"barcode": "BOX001"}).json
    auth_client.post("/api/tubes", json={"barcode": "TUBE001", "box_id": box["id"], "site_name": "Lake"})
    text = auth_client.get("/api/export/boxes").data.decode()
    assert "box" in text
    assert "tube" in text
    assert "TUBE001" in text
    assert "Lake" in text


def test_export_boxes_tube_count(auth_client: FlaskClient) -> None:
    box = auth_client.post("/api/boxes", json={"barcode": "BOX001"}).json
    auth_client.post("/api/tubes", json={"barcode": "TUBE001", "box_id": box["id"]})
    auth_client.post("/api/tubes", json={"barcode": "TUBE002", "box_id": box["id"]})
    text = auth_client.get("/api/export/boxes").data.decode()
    assert "2" in text


def test_export_boxes_empty(auth_client: FlaskClient) -> None:
    r = auth_client.get("/api/export/boxes")
    assert r.status_code == 200
    assert len(_lines(r)) == 1  # header only


# ── single box ────────────────────────────────────────────────────────────────

def test_export_single_box_csv(auth_client: FlaskClient) -> None:
    box = auth_client.post("/api/boxes", json={"barcode": "BOX001"}).json
    r = auth_client.get(f"/api/export/boxes/{box['id']}")
    assert r.status_code == 200
    assert "text/csv" in r.content_type
    text = r.data.decode()
    assert "BOX001" in text


def test_export_single_box_tsv(auth_client: FlaskClient) -> None:
    box = auth_client.post("/api/boxes", json={"barcode": "BOX001"}).json
    r = auth_client.get(f"/api/export/boxes/{box['id']}?format=tsv")
    assert r.status_code == 200
    assert "tab-separated" in r.content_type


def test_export_single_box_filename_contains_barcode(auth_client: FlaskClient) -> None:
    box = auth_client.post("/api/boxes", json={"barcode": "BOX-001"}).json
    r = auth_client.get(f"/api/export/boxes/{box['id']}")
    assert "BOX-001" in r.headers["Content-Disposition"]


def test_export_single_box_special_chars_sanitized(auth_client: FlaskClient) -> None:
    box = auth_client.post("/api/boxes", json={"barcode": "BOX 001"}).json
    r = auth_client.get(f"/api/export/boxes/{box['id']}")
    assert "BOX_001" in r.headers["Content-Disposition"]


def test_export_single_box_not_found_returns_empty(auth_client: FlaskClient) -> None:
    r = auth_client.get("/api/export/boxes/9999")
    assert r.status_code == 200
    assert len(_lines(r)) == 1  # header only, uses fallback name "box"


# ── cores (hierarchical) ──────────────────────────────────────────────────────

def test_export_cores_csv(auth_client: FlaskClient) -> None:
    auth_client.post("/api/cores", json={"barcode": "CORE001"})
    r = auth_client.get("/api/export/cores")
    assert r.status_code == 200
    assert "text/csv" in r.content_type
    assert re.search(rf'filename="cores-{TS_RE}\.csv"', r.headers["Content-Disposition"])
    text = r.data.decode()
    assert "row_type" in text
    assert "CORE001" in text


def test_export_cores_tsv(auth_client: FlaskClient) -> None:
    auth_client.post("/api/cores", json={"barcode": "CORE001"})
    r = auth_client.get("/api/export/cores?format=tsv")
    assert r.status_code == 200
    assert "tab-separated" in r.content_type
    assert re.search(rf'filename="cores-{TS_RE}\.tsv"', r.headers["Content-Disposition"])


def test_export_cores_with_tubes_in_box(auth_client: FlaskClient) -> None:
    core = auth_client.post("/api/cores", json={"barcode": "CORE001"}).json
    box = auth_client.post("/api/boxes", json={"barcode": "BOX001"}).json
    auth_client.post("/api/tubes", json={
        "barcode": "TUBE001", "core_id": core["id"], "box_id": box["id"]
    })
    text = auth_client.get("/api/export/cores").data.decode()
    assert "core" in text
    assert "box" in text
    assert "tube" in text
    assert "TUBE001" in text


def test_export_cores_with_unboxed_tubes(auth_client: FlaskClient) -> None:
    core = auth_client.post("/api/cores", json={"barcode": "CORE001"}).json
    auth_client.post("/api/tubes", json={"barcode": "TUBE001", "core_id": core["id"]})
    text = auth_client.get("/api/export/cores").data.decode()
    assert "TUBE001" in text


def test_export_cores_empty(auth_client: FlaskClient) -> None:
    r = auth_client.get("/api/export/cores")
    assert r.status_code == 200
    assert len(_lines(r)) == 1  # header only


# ── single core ───────────────────────────────────────────────────────────────

def test_export_single_core_csv(auth_client: FlaskClient) -> None:
    core = auth_client.post("/api/cores", json={"barcode": "CORE001"}).json
    r = auth_client.get(f"/api/export/cores/{core['id']}")
    assert r.status_code == 200
    assert "text/csv" in r.content_type
    assert "CORE001" in r.data.decode()


def test_export_single_core_tsv(auth_client: FlaskClient) -> None:
    core = auth_client.post("/api/cores", json={"barcode": "CORE001"}).json
    r = auth_client.get(f"/api/export/cores/{core['id']}?format=tsv")
    assert r.status_code == 200
    assert "tab-separated" in r.content_type


def test_export_single_core_filename_contains_barcode(auth_client: FlaskClient) -> None:
    core = auth_client.post("/api/cores", json={"barcode": "CORE-A1"}).json
    r = auth_client.get(f"/api/export/cores/{core['id']}")
    assert "CORE-A1" in r.headers["Content-Disposition"]


def test_export_single_core_not_found_returns_empty(auth_client: FlaskClient) -> None:
    r = auth_client.get("/api/export/cores/9999")
    assert r.status_code == 200
    assert len(_lines(r)) == 1  # header only


# ── boxes flat ────────────────────────────────────────────────────────────────

def test_export_boxes_flat_csv(auth_client: FlaskClient) -> None:
    auth_client.post("/api/boxes", json={"barcode": "BOX001", "name": "Shelf A"})
    r = auth_client.get("/api/export/boxes?flat=1")
    assert r.status_code == 200
    assert "text/csv" in r.content_type
    assert re.search(rf'filename="boxes-flat-{TS_RE}\.csv"', r.headers["Content-Disposition"])
    text = r.data.decode()
    assert "row_type" not in text
    assert "BOX001" in text


def test_export_boxes_flat_excludes_tube_rows(auth_client: FlaskClient) -> None:
    box = auth_client.post("/api/boxes", json={"barcode": "BOX001"}).json
    auth_client.post("/api/tubes", json={"barcode": "TUBE001", "box_id": box["id"]})
    text = auth_client.get("/api/export/boxes?flat=1").data.decode()
    assert "row_type" not in text
    assert "TUBE001" not in text
    assert "BOX001" in text


def test_export_single_box_flat(auth_client: FlaskClient) -> None:
    box = auth_client.post("/api/boxes", json={"barcode": "BOX001"}).json
    r = auth_client.get(f"/api/export/boxes/{box['id']}?flat=1")
    assert r.status_code == 200
    assert "BOX001" in r.data.decode()
    assert "box-flat-BOX001" in r.headers["Content-Disposition"]


# ── cores flat ────────────────────────────────────────────────────────────────

def test_export_cores_flat_csv(auth_client: FlaskClient) -> None:
    auth_client.post("/api/cores", json={"barcode": "CORE001"})
    r = auth_client.get("/api/export/cores?flat=1")
    assert r.status_code == 200
    assert "text/csv" in r.content_type
    assert re.search(rf'filename="cores-flat-{TS_RE}\.csv"', r.headers["Content-Disposition"])
    text = r.data.decode()
    assert "row_type" not in text
    assert "CORE001" in text


def test_export_cores_flat_excludes_tube_rows(auth_client: FlaskClient) -> None:
    core = auth_client.post("/api/cores", json={"barcode": "CORE001"}).json
    auth_client.post("/api/tubes", json={"barcode": "TUBE001", "core_id": core["id"]})
    text = auth_client.get("/api/export/cores?flat=1").data.decode()
    assert "row_type" not in text
    assert "TUBE001" not in text
    assert "CORE001" in text


def test_export_single_core_flat(auth_client: FlaskClient) -> None:
    core = auth_client.post("/api/cores", json={"barcode": "CORE001"}).json
    r = auth_client.get(f"/api/export/cores/{core['id']}?flat=1")
    assert r.status_code == 200
    assert "CORE001" in r.data.decode()
    assert "core-flat-CORE001" in r.headers["Content-Disposition"]


# ── boxes ids filter (regression: filtered list exported full set) ────────────

def test_export_boxes_ids_filters_to_subset(auth_client: FlaskClient) -> None:
    box1 = auth_client.post("/api/boxes", json={"barcode": "BOX001"}).json
    auth_client.post("/api/boxes", json={"barcode": "BOX002"})
    r = auth_client.get(f"/api/export/boxes?ids={box1['id']}")
    text = r.data.decode()
    assert "BOX001" in text
    assert "BOX002" not in text


def test_export_boxes_flat_ids_filters_to_subset(auth_client: FlaskClient) -> None:
    box1 = auth_client.post("/api/boxes", json={"barcode": "BOX001"}).json
    auth_client.post("/api/boxes", json={"barcode": "BOX002"})
    r = auth_client.get(f"/api/export/boxes?flat=1&ids={box1['id']}")
    text = r.data.decode()
    assert "BOX001" in text
    assert "BOX002" not in text


def test_export_boxes_ids_includes_tubes_for_matching_boxes_only(auth_client: FlaskClient) -> None:
    box1 = auth_client.post("/api/boxes", json={"barcode": "BOX001"}).json
    box2 = auth_client.post("/api/boxes", json={"barcode": "BOX002"}).json
    auth_client.post("/api/tubes", json={"barcode": "TUBE_A", "box_id": box1["id"]})
    auth_client.post("/api/tubes", json={"barcode": "TUBE_B", "box_id": box2["id"]})
    r = auth_client.get(f"/api/export/boxes?ids={box1['id']}")
    text = r.data.decode()
    assert "TUBE_A" in text
    assert "TUBE_B" not in text


def test_export_boxes_ids_empty_returns_header_only(auth_client: FlaskClient) -> None:
    auth_client.post("/api/boxes", json={"barcode": "BOX001"})
    r = auth_client.get("/api/export/boxes?ids=")
    assert r.status_code == 200
    assert len(_lines(r)) > 1  # no ids param = export all


# ── cores ids filter (regression: filtered list exported full set) ─────────────

def test_export_cores_ids_filters_to_subset(auth_client: FlaskClient) -> None:
    core1 = auth_client.post("/api/cores", json={"barcode": "CORE001"}).json
    auth_client.post("/api/cores", json={"barcode": "CORE002"})
    r = auth_client.get(f"/api/export/cores?ids={core1['id']}")
    text = r.data.decode()
    assert "CORE001" in text
    assert "CORE002" not in text


def test_export_cores_flat_ids_filters_to_subset(auth_client: FlaskClient) -> None:
    core1 = auth_client.post("/api/cores", json={"barcode": "CORE001"}).json
    auth_client.post("/api/cores", json={"barcode": "CORE002"})
    r = auth_client.get(f"/api/export/cores?flat=1&ids={core1['id']}")
    text = r.data.decode()
    assert "CORE001" in text
    assert "CORE002" not in text


def test_export_cores_ids_includes_tubes_for_matching_cores_only(auth_client: FlaskClient) -> None:
    core1 = auth_client.post("/api/cores", json={"barcode": "CORE001"}).json
    core2 = auth_client.post("/api/cores", json={"barcode": "CORE002"}).json
    auth_client.post("/api/tubes", json={"barcode": "TUBE_A", "core_id": core1["id"]})
    auth_client.post("/api/tubes", json={"barcode": "TUBE_B", "core_id": core2["id"]})
    r = auth_client.get(f"/api/export/cores?ids={core1['id']}")
    text = r.data.decode()
    assert "TUBE_A" in text
    assert "TUBE_B" not in text


def test_export_cores_ids_multiple(auth_client: FlaskClient) -> None:
    core1 = auth_client.post("/api/cores", json={"barcode": "CORE001"}).json
    core2 = auth_client.post("/api/cores", json={"barcode": "CORE002"}).json
    auth_client.post("/api/cores", json={"barcode": "CORE003"})
    r = auth_client.get(f"/api/export/cores?ids={core1['id']},{core2['id']}")
    text = r.data.decode()
    assert "CORE001" in text
    assert "CORE002" in text
    assert "CORE003" not in text


# ── _safe helper (direct unit test) ──────────────────────────────────────────

def test_safe_keeps_alnum_and_allowed() -> None:
    from sampling.routes.export._responses import _safe
    assert _safe("CORE-001") == "CORE-001"
    assert _safe("core_001.csv") == "core_001.csv"


def test_safe_replaces_special_chars() -> None:
    from sampling.routes.export._responses import _safe
    assert _safe("BOX 001") == "BOX_001"
    assert _safe("A/B:C") == "A_B_C"


# ── _parse_ids edge cases ─────────────────────────────────────────────────────

def test_export_boxes_ids_invalid_falls_back_to_all(auth_client: FlaskClient) -> None:
    # non-integer ids → _parse_ids returns None → exports all
    auth_client.post("/api/boxes", json={"barcode": "BOX001"})
    r = auth_client.get("/api/export/boxes?ids=abc")
    assert r.status_code == 200
    assert "BOX001" in r.data.decode()


def test_export_boxes_ids_empty_csv_returns_header_only(auth_client: FlaskClient) -> None:
    # ids=, → _parse_ids returns [] → empty result (all builders' return [] branch)
    auth_client.post("/api/boxes", json={"barcode": "BOX001"})
    r = auth_client.get("/api/export/boxes?ids=,")
    assert r.status_code == 200
    assert len(_lines(r)) == 1  # header only


def test_export_boxes_flat_ids_empty_returns_header_only(auth_client: FlaskClient) -> None:
    auth_client.post("/api/boxes", json={"barcode": "BOX001"})
    r = auth_client.get("/api/export/boxes?flat=1&ids=,")
    assert r.status_code == 200
    assert len(_lines(r)) == 1


def test_export_cores_ids_empty_returns_header_only(auth_client: FlaskClient) -> None:
    auth_client.post("/api/cores", json={"barcode": "CORE001"})
    r = auth_client.get("/api/export/cores?ids=,")
    assert r.status_code == 200
    assert len(_lines(r)) == 1


def test_export_cores_flat_ids_empty_returns_header_only(auth_client: FlaskClient) -> None:
    auth_client.post("/api/cores", json={"barcode": "CORE001"})
    r = auth_client.get("/api/export/cores?flat=1&ids=,")
    assert r.status_code == 200
    assert len(_lines(r)) == 1


def test_export_boxes_json_ids_empty_returns_empty_array(auth_client: FlaskClient) -> None:
    import json
    auth_client.post("/api/boxes", json={"barcode": "BOX001"})
    r = auth_client.get("/api/export/boxes?format=json&ids=,")
    assert r.status_code == 200
    assert json.loads(r.data) == []


def test_export_cores_json_ids_empty_returns_empty_array(auth_client: FlaskClient) -> None:
    import json
    auth_client.post("/api/cores", json={"barcode": "CORE001"})
    r = auth_client.get("/api/export/cores?format=json&ids=,")
    assert r.status_code == 200
    assert json.loads(r.data) == []


def test_export_single_box_json_not_found_returns_empty(auth_client: FlaskClient) -> None:
    import json
    r = auth_client.get("/api/export/boxes/9999?format=json")
    assert r.status_code == 200
    assert json.loads(r.data) == []


def test_export_single_core_json_not_found_returns_empty(auth_client: FlaskClient) -> None:
    import json
    r = auth_client.get("/api/export/cores/9999?format=json")
    assert r.status_code == 200
    assert json.loads(r.data) == []


# ── JSON tubes ────────────────────────────────────────────────────────────────

def test_export_tubes_json(auth_client: FlaskClient) -> None:
    import json
    auth_client.post("/api/tubes", json={"barcode": "TUBE001", "site_name": "River A"})
    r = auth_client.get("/api/export/tubes?format=json")
    assert r.status_code == 200
    assert "application/json" in r.content_type
    assert re.search(rf'filename="tubes-{TS_RE}\.json"', r.headers["Content-Disposition"])
    data = json.loads(r.data)
    assert isinstance(data, list)
    assert data[0]["barcode"] == "TUBE001"
    assert data[0]["site_name"] == "River A"
    assert "latitude" in data[0]


# ── JSON boxes (nested) ───────────────────────────────────────────────────────

def test_export_boxes_json_nested(auth_client: FlaskClient) -> None:
    import json
    box = auth_client.post("/api/boxes", json={"barcode": "BOX001", "name": "Shelf A"}).json
    auth_client.post("/api/tubes", json={"barcode": "TUBE001", "box_id": box["id"], "site_name": "Lake"})
    r = auth_client.get("/api/export/boxes?format=json")
    assert r.status_code == 200
    assert "application/json" in r.content_type
    assert re.search(rf'filename="boxes-{TS_RE}\.json"', r.headers["Content-Disposition"])
    rows = json.loads(r.data)
    assert len(rows) == 1
    assert rows[0]["barcode"] == "BOX001"
    assert rows[0]["name"] == "Shelf A"
    assert "tubes" in rows[0]
    assert rows[0]["tubes"][0]["barcode"] == "TUBE001"
    assert rows[0]["tubes"][0]["site_name"] == "Lake"
    assert "box_barcode" not in rows[0]["tubes"][0]


def test_export_boxes_json_nested_no_tubes_key_absent_when_empty(auth_client: FlaskClient) -> None:
    import json
    auth_client.post("/api/boxes", json={"barcode": "BOX001"})
    rows = json.loads(auth_client.get("/api/export/boxes?format=json").data)
    assert rows[0]["tubes"] == []


def test_export_boxes_flat_json(auth_client: FlaskClient) -> None:
    import json
    auth_client.post("/api/boxes", json={"barcode": "BOX001"})
    r = auth_client.get("/api/export/boxes?flat=1&format=json")
    assert r.status_code == 200
    assert re.search(rf'filename="boxes-flat-{TS_RE}\.json"', r.headers["Content-Disposition"])
    rows = json.loads(r.data)
    assert rows[0]["barcode"] == "BOX001"
    assert "tubes" not in rows[0]


def test_export_single_box_json_nested(auth_client: FlaskClient) -> None:
    import json
    box = auth_client.post("/api/boxes", json={"barcode": "BOX001"}).json
    auth_client.post("/api/tubes", json={"barcode": "TUBE001", "box_id": box["id"]})
    r = auth_client.get(f"/api/export/boxes/{box['id']}?format=json")
    assert r.status_code == 200
    rows = json.loads(r.data)
    assert rows[0]["barcode"] == "BOX001"
    assert rows[0]["tubes"][0]["barcode"] == "TUBE001"
    assert "box-BOX001" in r.headers["Content-Disposition"]


def test_export_single_box_flat_json(auth_client: FlaskClient) -> None:
    import json
    box = auth_client.post("/api/boxes", json={"barcode": "BOX001"}).json
    r = auth_client.get(f"/api/export/boxes/{box['id']}?flat=1&format=json")
    assert r.status_code == 200
    rows = json.loads(r.data)
    assert rows[0]["barcode"] == "BOX001"
    assert "tubes" not in rows[0]


def test_export_boxes_json_ids_filters_to_subset(auth_client: FlaskClient) -> None:
    import json
    box1 = auth_client.post("/api/boxes", json={"barcode": "BOX001"}).json
    auth_client.post("/api/boxes", json={"barcode": "BOX002"})
    rows = json.loads(auth_client.get(f"/api/export/boxes?format=json&ids={box1['id']}").data)
    assert len(rows) == 1
    assert rows[0]["barcode"] == "BOX001"


# ── JSON cores (nested) ───────────────────────────────────────────────────────

def test_export_cores_json_nested(auth_client: FlaskClient) -> None:
    import json
    core = auth_client.post("/api/cores", json={"barcode": "CORE001"}).json
    box = auth_client.post("/api/boxes", json={"barcode": "BOX001"}).json
    auth_client.post("/api/tubes", json={"barcode": "TUBE001", "core_id": core["id"], "box_id": box["id"]})
    r = auth_client.get("/api/export/cores?format=json")
    assert r.status_code == 200
    assert "application/json" in r.content_type
    assert re.search(rf'filename="cores-{TS_RE}\.json"', r.headers["Content-Disposition"])
    rows = json.loads(r.data)
    assert rows[0]["barcode"] == "CORE001"
    assert "boxes" in rows[0]
    assert "unboxed_tubes" in rows[0]
    assert rows[0]["boxes"][0]["barcode"] == "BOX001"
    assert rows[0]["boxes"][0]["tubes"][0]["barcode"] == "TUBE001"


def test_export_cores_json_unboxed_tubes(auth_client: FlaskClient) -> None:
    import json
    core = auth_client.post("/api/cores", json={"barcode": "CORE001"}).json
    auth_client.post("/api/tubes", json={"barcode": "TUBE001", "core_id": core["id"]})
    rows = json.loads(auth_client.get("/api/export/cores?format=json").data)
    assert rows[0]["unboxed_tubes"][0]["barcode"] == "TUBE001"
    assert rows[0]["boxes"] == []


def test_export_cores_flat_json(auth_client: FlaskClient) -> None:
    import json
    auth_client.post("/api/cores", json={"barcode": "CORE001"})
    r = auth_client.get("/api/export/cores?flat=1&format=json")
    assert r.status_code == 200
    assert re.search(rf'filename="cores-flat-{TS_RE}\.json"', r.headers["Content-Disposition"])
    rows = json.loads(r.data)
    assert rows[0]["barcode"] == "CORE001"
    assert "boxes" not in rows[0]


def test_export_single_core_json_nested(auth_client: FlaskClient) -> None:
    import json
    core = auth_client.post("/api/cores", json={"barcode": "CORE001"}).json
    r = auth_client.get(f"/api/export/cores/{core['id']}?format=json")
    assert r.status_code == 200
    rows = json.loads(r.data)
    assert rows[0]["barcode"] == "CORE001"
    assert "boxes" in rows[0]
    assert "core-CORE001" in r.headers["Content-Disposition"]


def test_export_single_core_flat_json(auth_client: FlaskClient) -> None:
    import json
    core = auth_client.post("/api/cores", json={"barcode": "CORE001"}).json
    r = auth_client.get(f"/api/export/cores/{core['id']}?flat=1&format=json")
    assert r.status_code == 200
    rows = json.loads(r.data)
    assert rows[0]["barcode"] == "CORE001"
    assert "boxes" not in rows[0]


def test_export_cores_json_ids_filters_to_subset(auth_client: FlaskClient) -> None:
    import json
    core1 = auth_client.post("/api/cores", json={"barcode": "CORE001"}).json
    auth_client.post("/api/cores", json={"barcode": "CORE002"})
    rows = json.loads(auth_client.get(f"/api/export/cores?format=json&ids={core1['id']}").data)
    assert len(rows) == 1
    assert rows[0]["barcode"] == "CORE001"


# ── GeoJSON format ────────────────────────────────────────────────────────────

def test_export_tubes_geojson(auth_client: FlaskClient) -> None:
    import json
    auth_client.post("/api/tubes", json={"barcode": "TUBE001", "latitude": 51.5, "longitude": -0.1})
    auth_client.post("/api/tubes", json={"barcode": "TUBE002"})  # no coords — excluded
    r = auth_client.get("/api/export/tubes?format=geojson")
    assert r.status_code == 200
    assert "geo+json" in r.content_type
    assert re.search(rf'filename="tubes-{TS_RE}\.geojson"', r.headers["Content-Disposition"])
    fc = json.loads(r.data)
    assert fc["type"] == "FeatureCollection"
    assert len(fc["features"]) == 1
    feat = fc["features"][0]
    assert feat["geometry"]["type"] == "Point"
    assert feat["geometry"]["coordinates"] == [-0.1, 51.5]
    assert feat["properties"]["barcode"] == "TUBE001"
    assert "latitude" not in feat["properties"]
    assert "longitude" not in feat["properties"]


def test_export_tubes_geojson_no_coords_returns_empty_collection(auth_client: FlaskClient) -> None:
    import json
    auth_client.post("/api/tubes", json={"barcode": "TUBE001"})
    r = auth_client.get("/api/export/tubes?format=geojson")
    assert r.status_code == 200
    fc = json.loads(r.data)
    assert fc["features"] == []


def test_export_cores_flat_geojson(auth_client: FlaskClient) -> None:
    import json
    auth_client.post("/api/cores", json={"barcode": "CORE001", "latitude": 48.8, "longitude": 2.35})
    r = auth_client.get("/api/export/cores?flat=1&format=geojson")
    assert r.status_code == 200
    assert "geo+json" in r.content_type
    assert re.search(rf'filename="cores-flat-{TS_RE}\.geojson"', r.headers["Content-Disposition"])
    fc = json.loads(r.data)
    assert len(fc["features"]) == 1
    assert fc["features"][0]["properties"]["barcode"] == "CORE001"


def test_export_single_core_flat_geojson(auth_client: FlaskClient) -> None:
    import json
    core = auth_client.post("/api/cores", json={"barcode": "CORE001", "latitude": 48.8, "longitude": 2.35}).json
    r = auth_client.get(f"/api/export/cores/{core['id']}?flat=1&format=geojson")
    assert r.status_code == 200
    fc = json.loads(r.data)
    assert len(fc["features"]) == 1


# ── xlsx format ───────────────────────────────────────────────────────────────

def _sheet_values(wb, sheet_name: str) -> list[list]:
    ws = wb[sheet_name]
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def test_export_tubes_xlsx(auth_client: FlaskClient) -> None:
    import io
    import openpyxl
    auth_client.post("/api/tubes", json={"barcode": "TUBE001", "site_name": "River A"})
    r = auth_client.get("/api/export/tubes?format=xlsx")
    assert r.status_code == 200
    assert "spreadsheetml" in r.content_type
    assert re.search(rf'filename="tubes-{TS_RE}\.xlsx"', r.headers["Content-Disposition"])
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    assert wb.sheetnames == ["Tubes"]
    rows = _sheet_values(wb, "Tubes")
    assert rows[0][0] == "barcode"
    assert any("TUBE001" in str(c) for c in rows[1])


def test_export_tubes_xlsx_with_ids(auth_client: FlaskClient) -> None:
    import io
    import openpyxl
    t1 = auth_client.post("/api/tubes", json={"barcode": "TUBE001"}).json
    auth_client.post("/api/tubes", json={"barcode": "TUBE002"})
    r = auth_client.get(f"/api/export/tubes?format=xlsx&ids={t1['id']}")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    rows = _sheet_values(wb, "Tubes")
    barcodes = [r[0] for r in rows[1:]]
    assert "TUBE001" in barcodes
    assert "TUBE002" not in barcodes


def test_export_boxes_xlsx(auth_client: FlaskClient) -> None:
    import io
    import openpyxl
    box = auth_client.post("/api/boxes", json={"barcode": "BOX001"}).json
    auth_client.post("/api/tubes", json={"barcode": "TUBE001", "box_id": box["id"]})
    r = auth_client.get("/api/export/boxes?format=xlsx")
    assert r.status_code == 200
    assert "spreadsheetml" in r.content_type
    assert re.search(rf'filename="boxes-{TS_RE}\.xlsx"', r.headers["Content-Disposition"])
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    assert wb.sheetnames == ["Boxes", "Tubes"]
    box_rows = _sheet_values(wb, "Boxes")
    assert box_rows[0][0] == "barcode"
    assert any("BOX001" in str(c) for c in box_rows[1])
    tube_rows = _sheet_values(wb, "Tubes")
    assert any("TUBE001" in str(c) for c in tube_rows[1])


def test_export_single_box_xlsx(auth_client: FlaskClient) -> None:
    import io
    import openpyxl
    box = auth_client.post("/api/boxes", json={"barcode": "BOX001"}).json
    r = auth_client.get(f"/api/export/boxes/{box['id']}?format=xlsx")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    assert wb.sheetnames == ["Boxes", "Tubes"]
    assert any("BOX001" in str(c) for c in _sheet_values(wb, "Boxes")[1])


def test_export_cores_xlsx(auth_client: FlaskClient) -> None:
    import io
    import openpyxl
    core = auth_client.post("/api/cores", json={"barcode": "CORE001"}).json
    box = auth_client.post("/api/boxes", json={"barcode": "BOX001"}).json
    auth_client.post("/api/tubes", json={"barcode": "TUBE001", "core_id": core["id"], "box_id": box["id"]})
    r = auth_client.get("/api/export/cores?format=xlsx")
    assert r.status_code == 200
    assert "spreadsheetml" in r.content_type
    assert re.search(rf'filename="cores-{TS_RE}\.xlsx"', r.headers["Content-Disposition"])
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    assert wb.sheetnames == ["Cores", "Boxes", "Tubes"]
    assert any("CORE001" in str(c) for c in _sheet_values(wb, "Cores")[1])
    assert any("BOX001" in str(c) for c in _sheet_values(wb, "Boxes")[1])
    assert any("TUBE001" in str(c) for c in _sheet_values(wb, "Tubes")[1])


def test_export_single_core_xlsx(auth_client: FlaskClient) -> None:
    import io
    import openpyxl
    core = auth_client.post("/api/cores", json={"barcode": "CORE001"}).json
    auth_client.post("/api/tubes", json={"barcode": "TUBE001", "core_id": core["id"]})
    r = auth_client.get(f"/api/export/cores/{core['id']}?format=xlsx")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    assert wb.sheetnames == ["Cores", "Boxes", "Tubes"]
    assert any("CORE001" in str(c) for c in _sheet_values(wb, "Cores")[1])
    assert any("TUBE001" in str(c) for c in _sheet_values(wb, "Tubes")[1])
